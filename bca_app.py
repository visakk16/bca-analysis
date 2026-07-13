import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import tempfile
import re
import base64
import io
import json

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
)

st.set_page_config(page_title='BCA Plate Analysis', layout='wide')


# ─── Stage 1: Parse the file and pull out raw standard absorbances ───────────

def parse_standards(absorbance_path: str, config_path: str = None, std_cols: list = None):
    """Reads the absorbance file and returns everything needed later, plus the
    raw (sorted) standard concentration/absorbance arrays for the user to edit."""
    df = pd.read_excel(absorbance_path, header=None)

    # --- Defaults ---
    conc_list = [2000, 1500, 1000, 750, 500, 250, 125, 25, 0]
    sample_names = []
    dilution_factor = 8.0
    loading_protein = 30.0
    loading_volume = None  # None = not provided by user

    # --- Optional config file ---
    if config_path:
        try:
            config_df = pd.read_excel(config_path)
            if 'Standard Concentration (\u00b5g/ml)' in config_df.columns:
                temp_conc = config_df['Standard Concentration (\u00b5g/ml)'].dropna().tolist()
                if temp_conc:
                    conc_list = temp_conc
            if 'Sample Names' in config_df.columns:
                temp_names = config_df['Sample Names'].dropna().tolist()
                if temp_names:
                    sample_names = temp_names
            if 'Parameter' in config_df.columns and 'Value' in config_df.columns:
                params_df = config_df[config_df['Parameter'].isin(
                    ['dilution_factor', 'loading_protein', 'loading_volume']
                )]
                for _, row in params_df.iterrows():
                    if row['Parameter'] == 'dilution_factor':
                        dilution_factor = float(row['Value'])
                    elif row['Parameter'] == 'loading_protein':
                        loading_protein = float(row['Value'])
                    elif row['Parameter'] == 'loading_volume':
                        loading_volume = float(row['Value'])
        except Exception:
            pass

    # --- Detect numeric header row ---
    header_like = False
    try:
        first_row_vals = df.iloc[0, 1:20].values.tolist()
        int_like = sum(
            1 for v in first_row_vals
            if re.match(r'^\s*\d+(?:\.0+)?\s*$', str(v).strip())
        )
        if int_like >= max(3, int(0.6 * len(first_row_vals))):
            header_like = True
    except Exception:
        pass

    try:
        first_data_label = str(df.iloc[1, 0]).strip()
        if len(first_data_label) == 1 and first_data_label.isalpha():
            header_like = True
    except Exception:
        pass

    base = 1 if header_like else 0

    if std_cols is None:
        std_cols = [1, 2]

    n_stds = len(conc_list)

    # --- Read standard absorbances ---
    try:
        if n_stds >= 9:
            std_block = df.iloc[base: base + 8, std_cols]
            physical_std_count = std_block.dropna(how='all').shape[0]
            standards = std_block.mean(axis=1).values
            if df.shape[1] > 4:
                try:
                    de_vals = pd.to_numeric(
                        pd.Series(df.iloc[base, [3, 4]].values), errors='coerce'
                    ).dropna()
                    de_val = float(de_vals.mean()) if not de_vals.empty else np.nan
                except Exception:
                    de_val = np.nan
                standards = np.concatenate([standards, [de_val]])
            else:
                standards = np.concatenate([standards, [np.nan]])
        else:
            std_block = df.iloc[base: base + n_stds, std_cols]
            physical_std_count = std_block.dropna(how='all').shape[0]
            standards = std_block.mean(axis=1).values

        nan_idx = np.where(np.isnan(standards))[0]
        if nan_idx.size > 0 and df.shape[1] > 4:
            for idx in nan_idx:
                row = base + int(idx) if idx < 8 else base
                try:
                    de_numeric = pd.to_numeric(
                        pd.Series(df.iloc[row, [3, 4]].values), errors='coerce'
                    )
                    if not de_numeric.dropna().empty:
                        standards[idx] = float(de_numeric.mean())
                except Exception:
                    pass
    except Exception as e:
        raise ValueError(f'Failed reading standard block: {e}')

    if len(standards) != n_stds or np.isnan(standards).any():
        missing = [i for i, v in enumerate(standards) if np.isnan(v)]
        raise ValueError(
            f'Expected {n_stds} standards, got valid values for '
            f'{n_stds - len(missing)}. Missing indices: {missing}. '
            f'Check the absorbance and config files and the chosen standard columns.'
        )

    conc_arr = np.array(conc_list, dtype=float)
    std_arr = np.array(standards, dtype=float)

    order = np.argsort(conc_arr)
    conc_sorted = conc_arr[order]
    std_sorted = std_arr[order]

    return {
        'df': df,
        'base': base,
        'physical_std_count': int(physical_std_count),
        'conc_sorted': conc_sorted,
        'std_sorted': std_sorted,
        'dilution_factor': dilution_factor,
        'loading_protein': loading_protein,
        'loading_volume': loading_volume,
        'sample_names': sample_names,
    }


# ─── Stage 2: Fit the (possibly edited) standards and compute sample volumes ──

def calculate_loading_info(ug_per_ul: float, target_ug: float, loading_volume, multiplier: float) -> dict:
    """Given one well's ug/ul, works out sample/buffer/scaled volumes and
    whether the target protein amount was actually hit.

    Returns a dict of plain numeric values (or None where not applicable) --
    no display formatting or 'N/A' strings live here. That keeps this function
    pure math, so it's easy to test and easy to trust on its own.
    """
    conc_valid = (not np.isnan(ug_per_ul)) and ug_per_ul > 0
    has_loading_volume = loading_volume is not None

    info = {'conc_valid': conc_valid, 'has_loading_volume': has_loading_volume}

    if not conc_valid:
        # Blank/empty well, or absorbance at/below the standard curve
        # intercept -- there's no usable concentration to compute a volume
        # from at all.
        return info

    sample_vol = float(target_ug) / ug_per_ul

    if not has_loading_volume:
        # No total loading volume set -- just report the raw sample volume,
        # there's no buffer/scaled/SDS concept without a lane total.
        info['sample_vol'] = sample_vol
        return info

    # Real concentration + a known lane total. Cap the sample volume at the
    # lane's total volume (can't pipette more sample than the lane holds);
    # if that cap kicks in, the target protein amount wasn't actually hit,
    # which we flag rather than silently reporting a "full" result.
    lv = float(loading_volume)
    sv = min(sample_vol, lv)
    buffer_vol = max(0.0, lv - sv)

    info.update({
        'sample_vol': sv,
        'buffer_vol': buffer_vol,
        'vol_nx': multiplier * sv,
        'buffer_vol_nx': multiplier * buffer_vol,
        'sds_dye': lv / 5.0,
        'actual_ug_loaded': sv * ug_per_ul,
        'under_loaded': sample_vol > lv,
    })
    return info


def format_result_row(abs_mean, conc, conc_with_dilution, ug_per_ul, loading_info: dict,
                       loading_volume, mx_label: str) -> dict:
    """Turns the raw numbers for one well into the display-ready row dict,
    handling all three scenarios (bad concentration / no loading volume set /
    normal case) in one place instead of three near-duplicate dictionaries."""
    row = {
        'Sample': None,
        'Absorbance': round(abs_mean, 3),
        'Concentration (ug/ml)': round(float(conc), 2) if not np.isnan(conc) else float('nan'),
        'ug/ml (with dilution)': round(float(conc_with_dilution), 2) if not np.isnan(conc_with_dilution) else float('nan'),
        'ug/ul': round(float(ug_per_ul), 4) if not np.isnan(ug_per_ul) else float('nan'),
    }

    conc_valid = loading_info['conc_valid']
    has_loading_volume = loading_info['has_loading_volume']

    if not conc_valid and has_loading_volume:
        row.update({
            'Sample Volume (ul)': 'N/A (conc <= 0)',
            'Buffer Volume (ul)': 'N/A (conc <= 0)',
            f'{mx_label} Sample Volume (ul)': 'N/A',
            f'{mx_label} Buffer Volume (ul)': 'N/A',
            '6X SDS Loading Dye (ul)': round(float(loading_volume) / 5.0, 2),
            'Target Met?': 'N/A (conc <= 0)',
            'Actual ug Loaded': 'N/A',
        })
    elif not has_loading_volume:
        row.update({
            'Sample Volume (ul)': round(loading_info['sample_vol'], 2) if conc_valid else 'N/A (conc <= 0)',
            'Buffer Volume (ul)': 'N/A',
            f'{mx_label} Sample Volume (ul)': 'N/A',
            f'{mx_label} Buffer Volume (ul)': 'N/A',
            '6X SDS Loading Dye (ul)': 'N/A',
            'Target Met?': 'N/A (no loading volume set)',
            'Actual ug Loaded': 'N/A',
        })
    else:
        # conc_valid and has_loading_volume -- the normal case
        row.update({
            'Sample Volume (ul)': round(loading_info['sample_vol'], 2),
            'Buffer Volume (ul)': round(loading_info['buffer_vol'], 2),
            f'{mx_label} Sample Volume (ul)': round(loading_info['vol_nx'], 2),
            f'{mx_label} Buffer Volume (ul)': round(loading_info['buffer_vol_nx'], 2),
            '6X SDS Loading Dye (ul)': round(loading_info['sds_dye'], 2),
            'Target Met?': 'No (under-loaded)' if loading_info['under_loaded'] else 'Yes',
            'Actual ug Loaded': round(loading_info['actual_ug_loaded'], 2),
        })

    return row


def run_analysis(
    df: pd.DataFrame,
    base: int,
    used_conc: np.ndarray,
    used_std: np.ndarray,
    all_conc: np.ndarray,
    all_std: np.ndarray,
    physical_std_count: int,
    dilution_factor: float,
    loading_volume,
    loading_protein: float,
    dilution_override: float = None,
    total_volume_override: float = None,
    round_digits: int = 6,
    target_ug_override: float = None,
    sample_names_override: list = None,
    multiplier: float = 2.0,
):
    # --- UI overrides ---
    if dilution_override is not None:
        try:
            dilution_factor = float(dilution_override)
        except Exception:
            pass

    if total_volume_override is not None:
        try:
            loading_volume = float(total_volume_override)
        except Exception:
            pass

    has_loading_volume = loading_volume is not None

    if used_conc.size < 2:
        raise ValueError('Not enough standards included to fit a line (need >= 2).')

    # --- Linear regression ---
    coeffs = np.polyfit(used_conc, used_std, 1)
    slope = float(coeffs[0])
    intercept = float(coeffs[1])

    rd = int(round_digits)
    slope_rounded = round(slope, rd)
    intercept_rounded = round(intercept, rd)

    preds = slope * used_conc + intercept
    ss_res = np.sum((used_std - preds) ** 2)
    ss_tot = np.sum((used_std - np.mean(used_std)) ** 2)
    r_squared = float(1.0 - ss_res / ss_tot) if ss_tot != 0 else 0.0

    # --- Standard curve figure ---
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.scatter(used_conc, used_std, color='blue', label='Standards used in fit')

    # Show excluded points (present in all_conc/all_std but not used_conc/used_std) in grey
    used_pairs = set(zip(np.round(used_conc, 6), np.round(used_std, 6)))
    excluded_mask = np.array([
        (round(c, 6), round(s, 6)) not in used_pairs for c, s in zip(all_conc, all_std)
    ])
    if excluded_mask.any():
        ax.scatter(
            all_conc[excluded_mask], all_std[excluded_mask],
            color='lightgray', label='Excluded points', marker='x'
        )

    x_line = np.linspace(float(np.min(used_conc)), float(np.max(used_conc)), 100)
    y_line = slope_rounded * x_line + intercept_rounded
    ax.plot(
        x_line, y_line, '--r',
        label=f'y = {slope_rounded:.{rd}f}x + {intercept_rounded:.{rd}f}\nR2 = {r_squared:.4f}'
    )
    for xc, yc in zip(all_conc, all_std):
        ax.annotate(
            f"{int(xc)}", (xc, yc),
            textcoords='offset points', xytext=(4, 4), fontsize=8, color='black'
        )
    x_padding = max(1.0, (float(np.max(all_conc)) - float(np.min(all_conc))) * 0.03)
    ax.set_xlim(float(np.min(all_conc)) - x_padding, float(np.max(all_conc)) + x_padding)
    ax.set_xlabel('Concentration (ug/ml)')
    ax.set_ylabel('Absorbance')
    ax.set_title('BCA Standard Curve')
    ax.legend()
    ax.grid(True, alpha=0.3)

    def calc_concentration(abs_value):
        return (abs_value - intercept_rounded) / slope_rounded

    def col_letter(idx: int) -> str:
        idx0 = int(idx)
        letters = ''
        while idx0 >= 0:
            letters = chr(ord('A') + (idx0 % 26)) + letters
            idx0 = idx0 // 26 - 1
        return letters

    # --- Multiplier label ---
    try:
        mx_label = f"{int(multiplier)}X" if float(multiplier).is_integer() else f"{multiplier}X"
    except Exception:
        mx_label = 'NX'

    # --- Read samples ---
    # Row A of the first sample column-pair (3,4) is occupied by Standard 9,
    # so that pair only has 7 usable sample rows (B-H). Every later column
    # pair has a completely empty row A available for real samples, so those
    # pairs should read all 8 rows (A-H), not skip row A too.
    sample_col_pairs = [(3, 4), (5, 6), (7, 8), (9, 10)]
    results = []
    sample_index = 0
    names_provided = bool(sample_names_override)

    target_ug = loading_protein if target_ug_override is None else float(target_ug_override)

    for pair_idx, (c1, c2) in enumerate(sample_col_pairs):
        if names_provided and sample_index >= len(sample_names_override):
            break
        if df.shape[1] <= max(c1, c2):
            continue

        sample_start = base + 1 if pair_idx == 0 else base
        sample_data = df.iloc[sample_start: sample_start + 8, [c1, c2]].dropna(how='all')
        if sample_data.empty:
            continue

        for row_idx, row_vals in sample_data.iterrows():
            vals = pd.to_numeric(pd.Series(row_vals.values), errors='coerce').dropna().values
            if vals.size == 0:
                continue

            abs_mean = float(np.mean(vals))
            conc = calc_concentration(abs_mean)
            conc_with_dilution = conc * dilution_factor
            ug_per_ul = conc_with_dilution / 1000.0

            loading_info = calculate_loading_info(ug_per_ul, target_ug, loading_volume, multiplier)
            row = format_result_row(
                abs_mean, conc, conc_with_dilution, ug_per_ul,
                loading_info, loading_volume, mx_label,
            )

            if sample_names_override and sample_index < len(sample_names_override):
                candidate = sample_names_override[sample_index].strip()
                if candidate == '':
                    sample_index += 1
                    continue
                sample_name = candidate
            elif sample_names_override is None and sample_index < 0:
                sample_name = ''
            else:
                sample_name = f"{col_letter(c1)}{int(row_idx) + 1}"

            row['Sample'] = sample_name
            results.append(row)
            sample_index += 1

    results_df = pd.DataFrame(results)
    if not results_df.empty:
        cols = ['Sample'] + [c for c in results_df.columns if c != 'Sample']
        results_df = results_df[cols]

    if sample_names_override:
        provided_nonempty = [s.strip() for s in sample_names_override if s and s.strip()]
        if provided_nonempty and not results_df.empty:
            results_df = results_df[results_df['Sample'].isin(set(provided_nonempty))].copy()

    stats_out = {
        'slope': slope,
        'intercept': intercept,
        'slope_rounded': slope_rounded,
        'intercept_rounded': intercept_rounded,
        'r_squared': r_squared,
        'conc_sorted': all_conc.tolist(),
        'std_sorted': all_std.tolist(),
        'used_conc': used_conc.tolist(),
        'used_std': used_std.tolist(),
        'physical_std_count': int(physical_std_count),
        'base': int(base),
        'total_volume': float(loading_volume) if has_loading_volume else None,
        'loading_protein': float(loading_protein),
    }

    return results_df, stats_out, fig


# ─── Stage 3: Build downloadable Excel / PDF reports ──────────────────────────

def _clean_cell(val):
    """Turn NaN floats into a plain 'N/A' string; leave everything else as-is."""
    if isinstance(val, float) and np.isnan(val):
        return 'N/A'
    return val


def build_excel_report(results_df: pd.DataFrame, stats_out: dict, fig, round_digits: int) -> bytes:
    """Builds an .xlsx with a Results sheet (summary stats + full results table)
    and a Standard Curve sheet (the chart, embedded as an image)."""
    rd = int(round_digits)
    wb = Workbook()

    # --- Results sheet ---
    ws = wb.active
    ws.title = 'Results'

    title_font = Font(name='Arial', bold=True, size=14)
    body_font = Font(name='Arial', size=10)
    header_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')

    ws['A1'] = 'BCA Plate Analysis Results'
    ws['A1'].font = title_font

    ws['A3'] = 'R2'
    ws['B3'] = round(stats_out['r_squared'], 4)
    ws['A4'] = 'Slope (rounded)'
    ws['B4'] = round(stats_out.get('slope_rounded', stats_out['slope']), rd)
    ws['A5'] = 'Intercept (rounded)'
    ws['B5'] = round(stats_out.get('intercept_rounded', stats_out['intercept']), rd)
    for r in (3, 4, 5):
        ws.cell(row=r, column=1).font = body_font
        ws.cell(row=r, column=2).font = body_font

    header_row = 7
    for c_idx, col_name in enumerate(results_df.columns, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for r_offset, row in enumerate(results_df.itertuples(index=False), start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=header_row + r_offset, column=c_idx, value=_clean_cell(val))
            cell.font = body_font
            cell.alignment = center

    for c_idx, col_name in enumerate(results_df.columns, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = max(14, len(str(col_name)) + 4)

    # --- Standard Curve sheet (chart image) ---
    ws_chart = wb.create_sheet('Standard Curve')
    img_buf = io.BytesIO()
    fig.savefig(img_buf, format='png', dpi=150, bbox_inches='tight')
    img_buf.seek(0)
    xl_img = XLImage(img_buf)
    ws_chart.add_image(xl_img, 'A1')

    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)
    return out_buf.getvalue()


def build_pdf_report(results_df: pd.DataFrame, stats_out: dict, fig, round_digits: int) -> bytes:
    """Builds a landscape-letter PDF with the summary stats, the standard curve
    chart, and the full results table."""
    rd = int(round_digits)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(letter),
        leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24,
    )
    styles = getSampleStyleSheet()

    header_cell_style = styles['Normal'].clone('header_cell')
    header_cell_style.fontSize = 6.5
    header_cell_style.leading = 8
    header_cell_style.textColor = colors.white
    header_cell_style.alignment = 1  # center

    body_cell_style = styles['Normal'].clone('body_cell')
    body_cell_style.fontSize = 6.5
    body_cell_style.leading = 8
    body_cell_style.alignment = 1  # center

    story = []
    story.append(Paragraph('BCA Plate Analysis Results', styles['Title']))
    story.append(Spacer(1, 10))
    story.append(Paragraph(f"R2: {stats_out['r_squared']:.4f}", styles['Normal']))
    story.append(Paragraph(
        f"Slope (rounded): {stats_out.get('slope_rounded', stats_out['slope']):.{rd}f}",
        styles['Normal'],
    ))
    story.append(Paragraph(
        f"Intercept (rounded): {stats_out.get('intercept_rounded', stats_out['intercept']):.{rd}f}",
        styles['Normal'],
    ))
    story.append(Spacer(1, 14))

    img_buf = io.BytesIO()
    fig.savefig(img_buf, format='png', dpi=150, bbox_inches='tight')
    img_buf.seek(0)
    story.append(RLImage(img_buf, width=6.0 * inch, height=3.6 * inch))
    story.append(Spacer(1, 14))

    table_data = [[Paragraph(str(c), header_cell_style) for c in results_df.columns]]
    for row in results_df.itertuples(index=False):
        table_data.append([
            Paragraph(str(_clean_cell(val)), body_cell_style) for val in row
        ])

    num_cols = len(results_df.columns)
    avail_width = landscape(letter)[0] - 48  # minus left/right margins
    col_width = avail_width / num_cols

    tbl = Table(table_data, colWidths=[col_width] * num_cols, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
    ]))
    story.append(tbl)

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


def build_plate_map_html() -> str:
    """Builds an HTML plate-map grid showing exactly how this app reads a
    plate: which wells are standards, which are samples, and which duplicate
    pairs belong together. This mirrors the app's real parsing logic (rows
    A-H, columns 1-10) -- it isn't just illustrative, it's the actual layout
    the code expects.

    Standards are shaded on a blue gradient (Standard 1 = darkest/highest
    concentration, Standard 9 = white/lowest -- mirroring how a real
    absorbance heatmap looks). Samples are plain white cells with black text.
    """
    rows = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    num_cols = 10
    num_standards = 9

    labels = {}
    for i, r in enumerate(rows):
        labels[(r, 1)] = f'Standard {i + 1}'
        labels[(r, 2)] = f'Standard {i + 1}'
    labels[('A', 3)] = 'Standard 9'
    labels[('A', 4)] = 'Standard 9'

    sample_num = 1
    sample_col_pairs = [(3, 4), (5, 6), (7, 8), (9, 10)]
    for pair_idx, (c1, c2) in enumerate(sample_col_pairs):
        # Row A of the first pair is occupied by Standard 9, so that pair
        # only has rows B-H available. Every later pair has a free row A,
        # so those pairs use all 8 rows (A-H).
        row_list = rows[1:] if pair_idx == 0 else rows
        for r in row_list:
            labels[(r, c1)] = f'Sample {sample_num}'
            labels[(r, c2)] = f'Sample {sample_num}'
            sample_num += 1

    def blue_shade(standard_num: int) -> str:
        """Standard 1 -> darkest blue, Standard 9 -> white. Linear
        interpolation between a dark blue and white across the 9 standards."""
        dark = (31, 95, 168)   # a strong blue
        light = (255, 255, 255)  # white
        t = (standard_num - 1) / (num_standards - 1)  # 0.0 .. 1.0
        r = round(dark[0] + t * (light[0] - dark[0]))
        g = round(dark[1] + t * (light[1] - dark[1]))
        b = round(dark[2] + t * (light[2] - dark[2]))
        return f'rgb({r},{g},{b})'

    header_cells = ''.join(f'<th style="padding:6px 10px;">{c}</th>' for c in range(1, num_cols + 1))
    body_rows = ''
    for r in rows:
        row_cells = f'<th style="padding:6px 10px;">{r}</th>'
        for c in range(1, num_cols + 1):
            label = labels.get((r, c), '')
            if label.startswith('Standard'):
                standard_num = int(label.split()[1])
                color = blue_shade(standard_num)
            else:
                color = '#FFFFFF'
            row_cells += (
                f'<td style="padding:6px 10px; text-align:center; color:#000000; '
                f'background:{color}; border:1px solid #999; font-size:12px;">{label}</td>'
            )
        body_rows += f'<tr>{row_cells}</tr>'

    html = f"""
    <table style="border-collapse:collapse; font-family:Arial, sans-serif;">
        <tr><th></th>{header_cells}</tr>
        {body_rows}
    </table>
    """
    return html


# ─── Streamlit UI ────────────────────────────────────────────────────────────

st.title('BCA Plate Analysis')
st.write('Upload an absorbance Excel file (.xlsx).')

with st.expander('Show example plate layout (click to expand)'):
    st.markdown(
        "This app reads plate wells in **duplicate pairs** -- each pair shares "
        "one absorbance value (matching labels below mark a pair, e.g. B3/B4). "
        "Standards fill columns 1-2 (rows A-H, one standard per row, highest "
        "concentration at Standard 1 down to lowest at Standard 8) plus a 9th "
        "standard at A3/A4 -- shaded darkest-to-white to mirror how a real "
        "absorbance heatmap looks, with Standard 9 being the lowest/blank. "
        "Once the standards are used up, sample loading starts right after -- "
        "**B3/B4 is Sample 1**. Row A of columns 3-4 is taken by Standard 9, "
        "so that first sample column-pair only has 7 rows (B-H) -- but every "
        "column-pair after that has a free row A, so those use all 8 rows "
        "(A-H). That gives 7 + 8 + 8 + 8 = **31 samples total** across the "
        "plate."
    )
    st.markdown(build_plate_map_html(), unsafe_allow_html=True)

abs_upl = st.file_uploader('Absorbance Excel file (.xlsx)', type=['xlsx'])

# Parse standards as soon as a (new) file is uploaded
if abs_upl is not None:
    if st.session_state.get('uploaded_file_name') != abs_upl.name:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as t_abs:
            t_abs.write(abs_upl.read())
            abs_path = t_abs.name

        st.session_state['uploaded_file_name'] = abs_upl.name
        st.session_state['abs_path'] = abs_path

        try:
            parsed = parse_standards(abs_path, config_path=None, std_cols=[1, 2])
            st.session_state['parsed'] = parsed
            st.session_state['std_editor_df'] = pd.DataFrame({
                'Concentration (ug/ml)': parsed['conc_sorted'],
                'Absorbance': parsed['std_sorted'],
                'Include in fit': [True] * len(parsed['conc_sorted']),
            })
            # clear any previous results since we have a new file
            for k in ('results_df', 'stats_out', 'fig'):
                st.session_state.pop(k, None)
        except Exception as e:
            st.error(f'Failed to parse standards: {e}')
            st.session_state.pop('parsed', None)
            st.session_state.pop('std_editor_df', None)

st.write('Optional: enter dilution factor if your samples were diluted (e.g., 5 for a 1:5 dilution).')
dilution_input = st.number_input('Dilution Factor:', min_value=1.0, value=5.0, step=1.0)

st.write(
    'Optional: choose decimal places to round slope and intercept '
    '(these rounded values are used to compute concentrations).'
)
round_digits = st.number_input(
    'Decimal places for slope/intercept:', min_value=0, max_value=10, value=6, step=1
)

st.write(
    'Enter target protein mass (ug). '
    'This is divided by the sample concentration to compute the required sample volume.'
)
target_ug_input = st.number_input('Target protein to load (ug):', min_value=0.1, value=20.0, step=0.1)

use_loading_volume = st.checkbox(
    'I have a total loading volume (required for buffer, scaled volume, and SDS dye columns)',
    value=True
)

if use_loading_volume:
    st.write('Enter total loading volume (ul) — the sum of sample + buffer per lane.')
    total_volume_input = st.number_input('Total loading volume (ul):', min_value=1.0, value=40.0, step=1.0)
else:
    total_volume_input = None
    st.info(
        'No loading volume — only Sample Volume will be calculated. '
        'Buffer, scaled, and SDS dye columns will show N/A.'
    )

st.write(
    'Optional: enter a multiplier for the scaled sample and buffer volume columns '
    '(e.g. 2 for 2X, 3 for 3X, 4 for 4X).'
)
multiplier_input = st.number_input('Volume multiplier:', min_value=1.0, value=2.0, step=1.0)

st.write(
    'Optional: paste sample names (one per line) to label samples instead of default well names. '
    'Leave a line blank to skip that position.'
)
sample_names_text = st.text_area('Sample names (one per line):', value='')
sample_names_list = sample_names_text.splitlines()

# ─── Editable standards table ────────────────────────────────────────────────

if 'std_editor_df' in st.session_state:
    st.subheader('Standard curve points')
    st.write(
        'Edit any concentration or absorbance value below, and check/uncheck any point(s) '
        'to include or exclude them from the linear fit.'
    )
    edited_std_df = st.data_editor(
        st.session_state['std_editor_df'],
        column_config={
            'Concentration (ug/ml)': st.column_config.NumberColumn('Concentration (ug/ml)', format='%.2f'),
            'Absorbance': st.column_config.NumberColumn('Absorbance', format='%.4f'),
            'Include in fit': st.column_config.CheckboxColumn('Include in fit'),
        },
        hide_index=True,
        num_rows='fixed',
        key='std_editor',
        use_container_width=True,
    )
    st.session_state['std_editor_df'] = edited_std_df

if st.button('Run analysis'):
    if 'parsed' not in st.session_state:
        st.error('Please upload an absorbance file first.')
    else:
        parsed = st.session_state['parsed']
        edited_df = st.session_state['std_editor_df']

        all_conc = edited_df['Concentration (ug/ml)'].astype(float).values
        all_std = edited_df['Absorbance'].astype(float).values
        mask = edited_df['Include in fit'].astype(bool).values
        used_conc = all_conc[mask]
        used_std = all_std[mask]

        if used_conc.size < 2:
            st.error('At least 2 standard points must be included to fit a line.')
        else:
            try:
                results_df, stats_out, fig = run_analysis(
                    df=parsed['df'],
                    base=parsed['base'],
                    used_conc=used_conc,
                    used_std=used_std,
                    all_conc=all_conc,
                    all_std=all_std,
                    physical_std_count=parsed['physical_std_count'],
                    dilution_factor=parsed['dilution_factor'],
                    loading_volume=parsed['loading_volume'],
                    loading_protein=parsed['loading_protein'],
                    dilution_override=dilution_input,
                    total_volume_override=total_volume_input,
                    round_digits=round_digits,
                    target_ug_override=target_ug_input,
                    sample_names_override=sample_names_list,
                    multiplier=multiplier_input,
                )
                st.session_state['results_df'] = results_df
                st.session_state['stats_out'] = stats_out
                st.session_state['fig'] = fig
                st.session_state['round_digits'] = round_digits
                st.session_state['multiplier'] = multiplier_input
            except Exception as e:
                st.error(f'Analysis failed: {e}')

# ─── Results ─────────────────────────────────────────────────────────────────

if 'results_df' in st.session_state:
    results_df = st.session_state['results_df']
    stats_out = st.session_state['stats_out']
    fig = st.session_state['fig']
    rd = int(st.session_state.get('round_digits', round_digits))

    st.metric('R2', f"{stats_out['r_squared']:.4f}")
    st.metric('Slope (rounded)', f"{stats_out.get('slope_rounded', stats_out['slope']):.{rd}f}")
    st.metric('Intercept (rounded)', f"{stats_out.get('intercept_rounded', stats_out['intercept']):.{rd}f}")

    st.pyplot(fig)
    st.dataframe(results_df)

    csv = results_df.to_csv(index=False).encode('utf-8')

    dl_col1, dl_col2, dl_col3 = st.columns(3)
    with dl_col1:
        st.download_button('Download results CSV', csv, file_name='bca_results.csv', mime='text/csv')
    with dl_col2:
        excel_bytes = build_excel_report(results_df, stats_out, fig, rd)
        st.download_button(
            'Download Excel report (.xlsx)',
            excel_bytes,
            file_name='bca_results.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    with dl_col3:
        pdf_bytes = build_pdf_report(results_df, stats_out, fig, rd)
        st.download_button(
            'Download PDF report (.pdf)',
            pdf_bytes,
            file_name='bca_results.pdf',
            mime='application/pdf',
        )

    # --- Print Chart & Table -------------------------------------------------
    # Streamlit's DOM/class names aren't stable enough to reliably hide
    # everything else on the page via print CSS, so instead we build a
    # standalone printable HTML document (chart as an embedded image, table
    # as plain HTML) and pop it open in a new window, then trigger the
    # browser's native print dialog on it.
    _img_buf = io.BytesIO()
    fig.savefig(_img_buf, format='png', dpi=150, bbox_inches='tight')
    _img_buf.seek(0)
    _img_b64 = base64.b64encode(_img_buf.read()).decode('utf-8')

    _table_html = results_df.to_html(index=False, border=0, na_rep='N/A')

    _print_doc = f"""
    <html>
    <head>
    <title>BCA Plate Analysis Results</title>
    <style>
        body {{ font-family: Arial, sans-serif; padding: 24px; color: #111; }}
        h2 {{ margin-top: 32px; }}
        img {{ max-width: 100%; height: auto; margin-bottom: 12px; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 8px; }}
        th, td {{ border: 1px solid #999; padding: 6px 10px; text-align: center; font-size: 12px; }}
        th {{ background: #eee; }}
        @media print {{
            body {{ padding: 0; }}
        }}
    </style>
    </head>
    <body>
        <h2>BCA Standard Curve</h2>
        <img src="data:image/png;base64,{_img_b64}" />
        <h2>Results</h2>
        {_table_html}
    </body>
    </html>
    """

    _print_doc_js = json.dumps(_print_doc)

    components.html(
        f"""
        <button
            onclick='
                var w = window.open("", "_blank");
                w.document.write({_print_doc_js});
                w.document.close();
                w.focus();
                setTimeout(function() {{ w.print(); }}, 500);
            '
            style="padding:8px 16px;font-size:14px;cursor:pointer;border-radius:6px;border:1px solid #999;background:#f5f5f5;"
        >
            🖨️ Print Chart &amp; Table
        </button>
        """,
        height=50,
    )

    if st.checkbox('Show raw standard arrays and input slice'):
        st.write('Concentrations (all, sorted):')
        st.write(stats_out.get('conc_sorted'))
        st.write('Absorbances (all, sorted):')
        st.write(stats_out.get('std_sorted'))
        st.write('Concentrations used for fit:')
        st.write(stats_out.get('used_conc'))
        st.write('Absorbances used for fit:')
        st.write(stats_out.get('used_std'))

        try:
            raw_df = pd.read_excel(st.session_state['abs_path'], header=None)
            st.write('Raw absorbance DataFrame (first 10 rows):')
            st.dataframe(raw_df.head(10))
            st.write('Wells used for standards (B&C, rows 0-7):')
            st.dataframe(raw_df.iloc[0:8, [1, 2]])
            st.write('Wells used for 9th standard (D2 & E2):')
            st.dataframe(pd.DataFrame([raw_df.iloc[0, [3, 4]]]))

            base_local = int(stats_out.get('base', 1))
            physical_std_count_local = int(stats_out.get('physical_std_count', len(stats_out.get('conc_sorted', []))))
            sample_start_local = base_local + physical_std_count_local
            st.write(
                f'Sample data slice start row (0-based): {sample_start_local} '
                f'-> Excel row {sample_start_local + 1}'
            )
            cols_to_show = [c for c in [3, 4, 5, 6, 7, 8, 9, 10] if c < raw_df.shape[1]]
            if cols_to_show:
                st.write('Sample area (D:K):')
                st.dataframe(raw_df.iloc[sample_start_local: sample_start_local + 20, cols_to_show])
            else:
                st.write('No sample columns found in the raw file.')
        except Exception as e:
            st.write('Could not read raw file for debug display:', e)

    try:
        std_df = pd.DataFrame({
            'Concentration_used_for_fit': stats_out.get('used_conc'),
            'Absorbance_used_for_fit': stats_out.get('used_std'),
        })
        csv_std = std_df.to_csv(index=False).encode('utf-8')
        st.download_button(
            'Download standards used for fit (CSV)',
            csv_std,
            file_name='standards_for_fit.csv',
            mime='text/csv',
        )
    except Exception:
        pass
